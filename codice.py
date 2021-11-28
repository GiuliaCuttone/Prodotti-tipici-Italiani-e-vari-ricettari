import openpyxl as ex
import pandas as pd
import os, csv
from rdflib import Graph, Literal, Namespace, URIRef
from rdflib.namespace import RDF

#regioni è un array che contiene i nomi dei fogli del file, che corrispondono ai nomi delle regioni
regioni = ex.load_workbook("Allegato_al_decreto_PAT_2019_prodotti_agroalimentari_tradizionali.xlsx").sheetnames 
regioni.remove("UMBRIA"); regioni.remove("PROV.TRENTO ")


#funzione che copia le righe di un foglio e le mette in una lista
def reader(regione):
    wb = ex.load_workbook("Allegato_al_decreto_PAT_2019_prodotti_agroalimentari_tradizionali.xlsx")
    sheet = wb[regione] #restituisce il foglio
    rows = []
    #itera per tutte le righe del foglio a partire dalla seconda e limitando la riga alla terza colonna
    for row in sheet.iter_rows(min_row=2, max_col=3):
        row_data = []
        for cell in row:
            #ogni posizione di row_data il valore di una cella
            if type(cell.value) == str:
                cell.value = cell.value.lower()
            row_data.append(cell.value)
        #Crea una lista di liste, ogni lista contiene i dati di una riga
        rows.append(row_data)
    wb.close()
    return rows

#per ogni foglio un file csv corrispondente ad una regione
for regione in regioni:
    rows = reader(regione) #funzione definita sopra
    wb = ex.Workbook() #crea nuovo workbook
    ws = wb.active #prende il foglio attivo
    for row in rows: #mette ogni riga nel nuovo workbook
        ws.append(row)

    ws['B1'].value = 'N'
    N = 2
    #dà i valori esatti a tutte le celle della colonna 'tipologia'
    #cicla finchè la cella della colonna B ha un valore, scorre verso il basso la colonna B
    while ws['B'+str(N)].value != None:
        #se la cella della colonna A ha un valore lo prende
        if ws['A'+str(N)].value != None:
            string = ws['A'+str(N)].value
        #se la cella della colonna A non ha un valore scrive quello precedentemente preso
        else:
            ws['A'+str(N)].value = string
        N = N + 1
    #una volta finito il ciclo prende la cella una posizione sotto della riga N nella colonna A e la cancella
    ws['A'+str(N+1)].value = None
    wb.save(regione+".xlsx")
    wb.close()
        
    df_xlsx = pd.read_excel(regione+'.xlsx')
    #aggiunge colonna regione e dà valori esatti per ogni riga
    df_xlsx['regione'] = str(regione).lower().strip()
    #rinomina le colonne
    df_xlsx = df_xlsx.rename(columns={'tipologia':'categoria'})
    #elimina la colonna denominata 'N'
    del df_xlsx['N'] 
    df_xlsx = df_xlsx.replace(' +','_',regex=True)
    df_xlsx = df_xlsx.replace('\W+','',regex=True)
    #da excel a csv     
    df_xlsx.to_csv(regione+'.csv', index = False,sep=',')
    #elimina il file xlsx
    os.remove(regione+".xlsx")

#unisce i file
df = pd.read_csv(regioni[0]+'.csv')
n = 1
while n < len(regioni):
    df = df.append(pd.read_csv(regioni[n]+'.csv'))
    n = n+1
df.to_csv('regioni_finale.csv', index = False, sep = ',')
#elimina i file singoli
for regione in regioni:
    os.remove(regione+'.csv')


df = pd.read_csv('TRENTINO.csv')
df1 = pd.read_csv('UMBRIA.csv',sep=';')
df2 = pd.read_csv('ricette.csv')

#pulizia file TRENTINO
#elimina la colonna denominata 'url'
del df['url']
#rinomina le colonne
df = df.rename(columns={'category': 'categoria', 
                        'DESCRIZIONE SINTETICA DEL PRODOTTO': 'descrizione', 
                        'CURIOSITA': 'curiosita', 
                        'METODICHE DI LAVORAZIONE E CONSERVAZIONE': 'metodiche_di_lavorazione_e_conservazione', 
                        'product_name': 'prodotto', 
                        'production_areas': 'area_produzione'})
#sostituisce spazi con underscore ed elimina caratteri non (letterali o numeri o underscore)
df = df.replace(' +','_',regex=True)
df = df.replace('\W+','',regex=True) 
#aggiunge colonna regione
df['regione'] = 'trentino'
#salva il file modificato
df.to_csv('TRENTINO_modificato.csv',index=False)


#pulizia file UMBRIA
#rinomina le colonne
df1 = df1.rename(columns={'Nome': 'prodotto',
                          'Metodiche di lavorazione conservazione e stagionatura': 'metodiche_di_lavorazione_e_conservazione',
                          'materiali e attrezzature per la preparazione': 'materiali_e_attrezzature_per_la_preparazione',
                          'Locali di lavorazione conservazione e stagionatura':'locali_di_lavorazione_conservazione_e_stagionatura',
                          'territorio interessato': 'area_produzione'})
#sostituisce spazi con underscore ed elimina caratteri non (letterali o numeri o underscore)
df1 = df1.replace(' +','_',regex=True)
df1 = df1.replace('\W+','',regex=True)
#aggiunge colonna regione
df1['regione'] = 'umbria'
#salva il file modificato
df1.to_csv('UMBRIA_modificato.csv',index=False)


#pulizia file ricette
#elimina la colonna denominata 'RecipeID'
del df2['RecipeID']
#rinomina le colonne
df2 = df2.rename(columns={'Title': 'nome', 
                        'Category': 'tipologia', 
                        'Ingredient': 'ingredienti', 
                        'Preparation': 'preparazione'})

#sostituisce spazi con underscore ed elimina caratteri non letterali o numeri o underscore
df2 = df2.replace(' +','_',regex=True)
df2 = df2.replace('\W+','',regex=True)
df2.to_csv('ricette_modificato.csv',index=False)



#costruisce il grafo rdf
g = Graph()         
pt = Namespace("http://www.prodotti_tipici.org/ontology/")
base_uri = "http://www.prodotti_tipici.org/resource/"
g.bind("pt", pt)

#dizionario
d = {'abruzzo': "http://dbpedia.org/resource/Abruzzo", 'basilicata': "http://dbpedia.org/resource/Basilicata", 
     'calabria': "http://dbpedia.org/resource/Calabria", 'emilia_romagna': "http://dbpedia.org/resource/Emilia-Romagna",
     'friuli': "http://dbpedia.org/resource/Friuli", 'lazio': "http://dbpedia.org/resource/Lazio",
     'liguria': "http://dbpedia.org/resource/Liguria", 'lombardia': "http://dbpedia.org/resource/Lombardia",
     'marche': "http://dbpedia.org/resource/Marche", 'molise': "http://dbpedia.org/resource/Molise",
     'piemonte': "http://dbpedia.org/resource/Piemonte", 'puglia': "http://dbpedia.org/resource/Puglia",
     'sardegna': "http://dbpedia.org/resource/Sardegna", 'sicilia': "http://dbpedia.org/resource/Sicilia",
     'toscana': "http://dbpedia.org/resource/Toscana", 'umbria': "http://dbpedia.org/resource/Umbria",
     'valle_daosta': "http://dbpedia.org/resource/Valle_d'Aosta", 'veneto': "http://dbpedia.org/resource/Veneto",
     'prov_bolzano': "http://dbpedia.org/resource/Bolzano", 'trentino': "http://dbpedia.org/resource/Trento",
     'campania': "http://dbpedia.org/resource/Campania"}

#aggiunge triple file regioni_finale
with open('regioni_finale.csv') as csvfile:
    lettore = csv.DictReader(csvfile)
    for row in lettore:
        uri_prodotto = base_uri+str(row['prodotto'][0].upper()+row['prodotto'][1:])#mette il primo carattere della stringa maiuscolo
        g.add([URIRef(uri_prodotto), RDF.type, pt.Prodotto])
        g.add([URIRef(uri_prodotto), pt.categoria, Literal(row['categoria'])])
        g.add([URIRef(uri_prodotto), pt.regione, URIRef(d[row['regione']])])#interlinking dbpedia

df_unito = pd.read_csv('TRENTINO_modificato.csv').merge(pd.read_csv('ricette_modificato.csv'), left_on='prodotto', right_on='nome')
df_unito.to_csv('ricette-prodotto.csv', index=False)

#aggiunge triple file trentino, file ricette e file che unisce le ricette ai prodotti
with open('TRENTINO_modificato.csv') as csvfile1, open('ricette_modificato.csv') as csvfile2, open('ricette-prodotto.csv') as csvfile3:
    lettore1 = csv.DictReader(csvfile1)
    lettore2 = csv.DictReader(csvfile2)
    lettore3 = csv.DictReader(csvfile3)
    for row in lettore1:
        uri_prodotto = base_uri+str(row['prodotto'])
        g.add([URIRef(uri_prodotto), RDF.type, pt.Prodotto])
        g.add([URIRef(uri_prodotto), pt.categoria, Literal(row['categoria'].lower())])
        g.add([URIRef(uri_prodotto), pt.regione, URIRef(d[row['regione']])])#interlinking dbpedia
        g.add([URIRef(uri_prodotto), pt.descrizione, Literal(row['descrizione'].lower())])
        g.add([URIRef(uri_prodotto), pt.curiosita, Literal(row['curiosita'].lower())])
        g.add([URIRef(uri_prodotto), pt.metodiche_lav_e_cons, Literal(row['metodiche_di_lavorazione_e_conservazione'].lower())])
        g.add([URIRef(uri_prodotto), pt.area_produzione, Literal(row['area_produzione'].lower())])
    for row in lettore2:
        uri_ricetta = base_uri+str(row['nome']+'_')
        g.add([URIRef(uri_ricetta), RDF.type, pt.Ricetta])
        g.add([URIRef(uri_ricetta), pt.tipologia, Literal(row['tipologia'].lower())])
        g.add([URIRef(uri_ricetta), pt.ingredienti, Literal(row['ingredienti'].lower())])        
        g.add([URIRef(uri_ricetta), pt.preparazione, Literal(row['preparazione'].lower())])
    for row in lettore3:
        uri_prodotto = base_uri+str(row['prodotto'])
        uri_ricetta = base_uri+str(row['nome']+'_')
        g.add([URIRef(uri_prodotto), pt.ha_ricetta, URIRef(uri_ricetta)])
        g.add([URIRef(uri_ricetta), pt.ha_prodotto, URIRef(uri_prodotto)])

#aggiunge triple file Umbria
with open('UMBRIA_modificato.csv') as csvfile:
    lettore = csv.DictReader(csvfile)
    for row in lettore:
        uri_prodotto = base_uri+str(row['prodotto'][0].upper()+row['prodotto'][1:].lower())#mette il primo carattere della stringa maiuscolo e il resto minuscolo
        g.add([URIRef(uri_prodotto), RDF.type, pt.Prodotto])
        g.add([URIRef(uri_prodotto), pt.categoria, Literal(row['categoria'].lower())])
        g.add([URIRef(uri_prodotto), pt.regione, URIRef(d[row['regione']])])#interlinking dbpedia
        g.add([URIRef(uri_prodotto), pt.descrizione, Literal(row['descrizione'].lower())])
        g.add([URIRef(uri_prodotto), pt.materiali_preparazione, Literal(row['materiali_e_attrezzature_per_la_preparazione'].lower())])
        g.add([URIRef(uri_prodotto), pt.metodiche_lav_e_cons, Literal(row['metodiche_di_lavorazione_e_conservazione'].lower())])
        g.add([URIRef(uri_prodotto), pt.area_produzione, Literal(row['area_produzione'].lower())])
        g.add([URIRef(uri_prodotto), pt.locali_lavorazione, Literal(row['locali_di_lavorazione_conservazione_e_stagionatura'].lower())])

g.serialize(destination='prodotti_tipici_rdf.ttl', format='turtle')   